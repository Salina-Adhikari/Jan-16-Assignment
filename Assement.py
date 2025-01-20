

from collections import deque
tasks=deque(['Do Homework','Wash Dishes','Iron Clothes'])
tasks_1=tasks.append('Make Dinner')
print(tasks)
tasks_2=tasks.remove('Iron Clothes')
print(tasks)
tasks.rotate(2)
print(tasks)

from collections import Counter
about_Counter="hi i am Salina i like food .hi nice to meet you.My friend like to eat food.Do you like to eat food"
word_counter = Counter(about_Counter.split())
print(word_counter)
print(word_counter.most_common(3))

from collections import defaultdict


product_inventory = defaultdict(list)

product_inventory["laptop"].extend(["dell", "acer"])
product_inventory["Phone"].extend(["oppo", "popo"])
product_inventory["dishes"].extend(["cup", "mug"])


def add_product(category, product):
    product_inventory[category].append(product)


add_product("Electronics", "tablet")
add_product("Phone", "samsung")
add_product("Books", "novel")    


for category, products in product_inventory.items():
    print(f"{category}: {', '.join(products)}")

import numpy as np
one_arrry=np.array([2,3,4,5])
print(np.mean(one_arrry))
print(np.sum(one_arrry))
print(np.std(one_arrry))
multiply=one_arrry*2
print(multiply)

First=np.array([[3,4],[9,4]])
Second=np.array([[6,7],[4,6]])
total=np.dot(First,Second)
print(total)


import csv
file_path="employeee_details.csv"

def create_csv():
    headers=["Name","Age","Department"]
    with open(file_path,mode="w",newline="")as file:
        writer=csv.writer(file)
        writer.writerow(headers)
    print("record has been added")

def add_details(name,age,department):
    with open(file_path,mode="a",newline="") as file:
        writer=csv.writer(file)
        writer.writerow([name,age,department])

    
def display_name():
     with open(file_path,mode="r",newline="") as file:
        reader=csv.reader(file)
        next(reader)
        for row in reader:
            if row:
                name,age,department=row
                if int(age)>30:
                   print(f"Name:{name},Age:{age}")






create_csv()
add_details("Roshan Shrestha",66,"IT department")
add_details("Ashmita Shresth",67,"Marketing")
add_details("rashi pun",80,"Production")
add_details("Sam Lun",27,"Human Resource")

display_name()

import openpyxl
file_path="Student_grades.xlsx"
def create_database(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Student_Data"

    sheet.append(["Name","Grade"])
    sheet.append(["Salina Adhikari",4.0])
    sheet.append(["Roshna Shrestha",4.0])
    sheet.append(["li pun",2.85])
    sheet.append(["Elija Muster",3.65 ])
    workbook.save(file_path)
    print("Student Data has been added")

def create_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
   

    print(f"Data in the sheet '{sheet.title}':")
    for row in sheet.iter_rows(values_only=True):
        print(row)



file_path="Data_Student.xlsx"
create_database(file_path)
create_excel(file_path)

import requests

url = "https://newsapi.org/v2/top-headlines?country=us&category=business&apiKey=c6815e64bc3e41a9a329019158379975"
response = requests.get(url)

if response.status_code == 200:
    data = response.json()  
    articles = data.get("articles", []) 
    for article in articles[:5]:  
        title = article.get("title", "No title")  
        print(f"Title: {title}")
else:
    print("Status code is 400")

import requests

url = "https://jsonplaceholder.typicode.com/posts"

post_data = {
    "title": "My First Post",
    "body": "This is the content of my first post.",
    "Author": "Salina Adhikari"
}
response = requests.post(url, json=post_data)

if response.status_code == 201:  
    print("Post created successfully!")
    data = response.json()
    print(f"Response Data: {data}")
else:
    print(f"Failed to create post. Status Code: {response.status_code}")
    print(f"Response: {response.text}")


import random
def SPR_Game():
    choices=["Scissors","Rock","Paper"]
    user=input("enter your choic(Scissors,Rock,Paper): ").capitalize()
    if user not in choices:
     print("Invalid choices")
    else:
     computer = random.choice(choices)
     print(f"The computer chose: {computer}")

    if user == computer:
     print("It's a tie!")
    elif (user == "Scissors" and computer == "Rock") or (user == "Paper" and computer == "Scissors"):
     print("Computer wins!")
    elif (user == "Rock" and computer == "Scissors") or (user == "Scissors" and computer == "Paper"):
     print("You win!")

SPR_Game()
import random

uppercase = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
lowercase = "abcdefghijklmnopqrstuvwxyz"
numbers = "0123456789"
special = "!@#$%^&*()-_=+[]{}|;:',.<>?/"
all_characters=uppercase+lowercase+numbers+special
user=int(input("Enter the length of password you want:"))
password = "".join(random.choice(all_characters) for _ in range(user))
print(f"Your new password is : {password}")
